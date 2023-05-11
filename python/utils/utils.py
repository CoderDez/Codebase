from datetime import date, datetime, timedelta
from io import BytesIO
from zipfile import ZipFile
from typing import Literal, Union
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pdfkit
import psycopg2
import json

class ValidatorUtility():
    
    def is_password_valid(val: str):
        """Returns True if password meets the following criteria:
        
        - between 8 and 20 characters in length.
        - at least 1 upper case char 
        - at least 1 lower case char
        - at least 1 number
        
        ,else False."""

        try:
            if len(val) >= 8 and len(val) <= 20:
                if any(s.isupper() for s in val):
                    if any(s.islower() for s in val):
                        if any(s.isdigit() for s in val):
                            return True
        except:
            pass

        return False
    
    def is_float(element) -> bool:
        """
        validates whether element is of type float.

        returns True if element can be converted to type float, else False.
        """
        try:
            float(element)
            return True
        except:
            return False
    

class DateUtility():

    def get_date_range(self, start: Union[str, datetime], end: Union[str, datetime]) -> list[str]:
        """returns a list of dates between start and end inclusive.
        
        format of returned str-represented dates: `YYYY-mm-dd`
        """
        try:
            if isinstance(start, str): start = self.get_date_from_string(start)
            if isinstance(end, str): end = self.get_date_from_string(end)

            date_range = []

            while (start <= end):
                date_range.append(self.get_string_from_date(start))
                start = self.add_days_to_date(start, 1)

            return date_range

        except:
            pass

    def add_days_to_date(self, date: datetime, days_add: int) -> datetime:
        """
        function to add days to a datetime object.
        
        returns date + `days_add` number of days 
        """
        return (date + timedelta(days_add))

    def get_week_start_and_end_dates(self, date: date) -> dict:
        """
        returns a dict containing start and end date for the argument `date` is in.

        format of `returning dict`:
        {
            `start_date`: str represented start date of week,

            `end_date`: str represented end date of week
        }
        """

        day_num = date.weekday()
        end_day_num = 6 - day_num
        start_date = date - timedelta(days=day_num)
        end_date = date + timedelta(days=end_day_num)

        return {
            'start_date': start_date.strftime("%Y-%m-%d"),
            'end_date': end_date.strftime("%Y-%m-%d")
        }
    
    def get_week_headers(self, date: date) -> list[str]:
        """
        returns a list of headers for a given week.

        header format: `day date/month` e.g `Mon 01/05`.
        """
        try:
            headers = []

            date_obj = self.get_week_start_and_end_dates(date)
            start_date = self.get_date_from_string(date_obj["start_date"])
            end_date = self.get_date_from_string(date_obj["end_date"])

            while (start_date <= end_date):
                header = "{} {}".format(start_date.strftime('%a'), start_date.strftime("%d/%m"))
                headers.append(header)
                start_date = self.add_days_to_date(start_date, 1)

            return headers
        
        except:
            pass

    def get_date_from_string(self, date_str: str)-> date:
        """creates and returns a date object generated from date_str.
        
        expected formats:

        - YYYY-mm-dd
        - dd-mm-YYYY"""

        comps = date_str.split("-")
        month = int(comps[1])
        year = day = None

        if len(comps[0]) == 4:
            year = int(comps[0])
            day = int(comps[2])
        else:
            year = int(comps[2])
            day = int(comps[0])

        return date(year, month, day)

    def get_string_from_date(self, dateobj: date) -> str:
        """
        returns a date represented as a string.

        format of returned string: `YYYY-mm-dd`."""

        return dateobj.strftime("%Y-%m-%d")

    def get_week_number(self, value: date) -> int:
        """returns the week number that value resides in."""
        return value.isocalendar()[1]
    

class ZipUtility():

    def populate_zip_files(self, files:dict) -> BytesIO:
        """
        creates a zip file using `files`.

        expects files to be in the following format:

        `file name` : {
            'data' : {
                'file_ext': str,

                'content' : str
            }
        }
        """
        content = BytesIO()
        with ZipFile(content, "w") as zipObj:
            for name, data in files.items():
                zipObj.writestr(name + data["file_extension"], data["content"])
        return content.getvalue()
    

class ExcelUtility():

    def __init__(self):
        self.work_b = Workbook()
        self.work_s = self.wb.active
        self.__setup()


    def merge_cells(self, start_col: int, start_row: int, end_col: int, end_row: int) -> bool:
        """
        merges cells in the active work sheet(work_s).

        returns True if cells successfully merged, else False."""
        try:
            self.work_s.merge_cells(
                start_column=start_col, start_row=start_row,
                end_column=end_col, end_row=end_row)
            return True
        except:
            return False
        
    def configure_dimensions(self, col: str, width: int) -> bool:
        """
        configures dimensions in the active worksheet (work_s)

        expects col to be an alphabethic character and width to be an int.

        returns True if configured successfully, else False."""
        try:
            col = col.capitalize()
            self.work_s.column_dimensions[col].width=width
            return True
        except:
            return False
        
    def configure_header(self, value: str, row: int, col: int, is_bold: bool, alignment: Literal["left", "center", "right"] = "left") -> bool:
        """
        configures a header in the active worksheet (work_s).

        returns True if configured successfully, else False.
        """
        try:
            cell = self.work_s.cell(row=row, column=col)
            cell.value=value
            if is_bold:
                bold=Font(bold=True)
                cell.font=bold
            align = Alignment(horizontal=alignment)
            cell.alignment = align
            return True
        except:
            return False
        
    def write(self, value: str, row: int, col: int, is_bold: bool, color: Union[None,str] = None, alignment: Literal["left", "center", "right"] = "left") -> bool:
        """
        writes content to the active worksheet (work_s).

        returns True if written successfully, else False.
        """

        try:
            if isinstance(value, list):
                value = ",".join(val for val in value)

            cell = self.work_s.cell(row=row, column=col)
            cell.value= self.value_formatter(value)

            align = Alignment(horizontal=alignment)
            cell.alignment=align

            if is_bold:
                bold = Font(bold=True)
                cell.font=bold

            if color:
                try:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                except:
                    pass

            return True
        
        except:
            return False
        
    def value_formatter(self, value) -> Union[float, str]:
        """formats a value that is to be written."""
        try:
            if isinstance(value,float) and not isinstance(value, int):
                return "{0:.2f}".format(value)
            else:
                return value
        except:
            return value


    def get_virtual_workbook(self):
        """returns an in-memory version of the workbook(work_b)."""
        try:
            with BytesIO() as buffer:
                self.wb.save(buffer)
                content = buffer.getvalue()   
                return content     
        except:
            pass

    
    def __setup__(self):
        # methods to 'setup' the excel file are called here.
        # these methods vary depending on requirements.
        pass


class PdfUtility():

    def __init__(self):
        self.validation = ValidatorUtility()
    

    def create_html_table(self, id: str, headers: list[str], tbody_rows: str, table_classes: list[str] = []):        
        html = f"""
        <table border="1" cellpadding="0" cellspacing="0" id="{id}" class={' '.join(cl for cl in table_classes)}>
            <thead>
                <tr>
                    {" ".join('<th>' + th + '</th>' for th in headers)}
                </tr>
            <thead>
            <tbody>
                {tbody_rows}
            </tbody>
        </table>
        """

        return html

    def generate_records(self, data: list[list], euro_indexes: list[int] = [], empty_record_td_count: int = 0) -> str:
        content = ""
        if isinstance(data, list):
            for r in data:
                tr = "<tr>"

                count = 0 

                for d in r:
                    td = "<td>"

                    if count in euro_indexes:
                        td += "&euro;"

                    if self.validation.is_float(d):
                        td += f"{float(d):.2f}</td>"

                    elif isinstance(d, list):
                        for i in d:
                            td += f"<span>{i}</span><br/><br/>"
                        td += "</td>"

                    else:
                        td += f"{d}</td>"

                    tr += td
                    count += 1

                tr += "</tr>"

                content += tr

        if content == "":
            content += "<tr>"
            for i in range(0, empty_record_td_count):
                content += "<td></td>"
            content += "</tr>"
        return content

    def get_pdf_css(self, file_path: str):
        """reads and returns css from .css file in file_path."""
        css = ""
        with open(file_path) as file:
            css = file.read()
        return css
    
    def create_html_document(self, css: str, body_content: str):
        html = f"""
        <!DOCTYPE html>
        <html>
            <head>
                <meta charset="utf-8"/>
                <style>{css}</style>
            </head>
            <body>
                {body_content}
            </body>
        </html>
        """
        return html
    
    def convert_html_string_to_pdf(self, string: str):
        path = '/bin/wkhtmltopdf'
        config = pdfkit.configuration(wkhtmltopdf=path)
        options = {'enable-local-file-access': None}
        data = pdfkit.from_string(string, output_path=None, configuration=config, options=options)
        return data


class PsycopgUtility():
    
    def __init__(self, user: str, password: str, host: str, port: str, database: str, auto_commit: bool = True):
        self.USER = user
        self.PASSWORD = password
        self.HOST = host
        self.PORT = port
        self.DB = database

        self.conn = psycopg2.connect(
            user=self.USER, password=self.PASSWORD,
            database = self.DB, host=self.HOST, port=self.PORT)
        
        self.conn.autocommit = auto_commit
        self.curs = self.conn.cursor()
        self.dict_cur = self.conn.cursor(dictionary=True)
        
    def __del__(self):
        self.conn.close()

    def delete_from(self, table: str, where_clause: str):
        try:
            sql = f"""
            DELETE FROM {table}
            {where_clause};
            """
            return self.executer(sql)
        except:
            return False
        
    def value_formatter(self, value):
        try:
            if isinstance(value, str):
                return f"'{value}'"

            elif isinstance(value, int) or isinstance(value, float) or isinstance(value, bool):
                return f"{value}"

            elif isinstance(value, date):
                return f"'{value.strftime('YYYY-mm-dd')}'"

            elif isinstance(value, datetime):
                return f"'{value.strftime('YYYY-mm-dd %H:%M:%S')}'"

            elif isinstance(value, bool):
                return f'{str(value)}'

            elif value is None:
                return "NULL"

            else:
                return value
        except:
            pass

    def create_insert_stmt(self, table: str, recs: list[list], cols: list[str] = []) -> str:
        try:
            sql = f"INSERT INTO {table} "

            if cols:
                sql += f" ( {','.join(col for col in cols)} ) "

            sql += f"VALUES "

            for r in recs:
                value = "(" + ",".join(self.value_formatter(d) for d in r) + "),"
                sql += value

            sql = sql[:-1] + ";"
            return sql
        except:
            pass
    
    def create_update_stmt(self, table: str, update_info: dict, where_clause: str = None):
        try:
            sql = f"""UPDATE {table} SET """
            sql += ",".join(
                f"{col} = {self.value_formatter(new_val)}" for col, new_val in update_info.items()
            )
            if where_clause:
                sql += f" {where_clause}"
            sql += ";"
            return sql
        except:
            pass

    def executer(self, sql: str, fetch_results: bool = False):
        try:
            self.curs.execute(sql)
            if fetch_results:
                return self.curs.fetchall()
            else:
                return True
        except:
            return False
        
    def create_tables_from_json(self, file_path: str, output_path: str, schema: str, non_null_fields: list[str], add_geom_column: bool = True):
        """
        creates SQL statements for the creation of tables.
        
        expects the json file to be in the following format:

        {
            "crs_id": integer,
            "Layers" : [
                {
                    'name': str,

                    'fields': [
                        {field_name: datatype}
                    ]
                },
            ]
        }

        ...

        Arguments 
        ---------
        `file_path`: path to json file.

        `output_path`: where the script is to be outputted.

        `schema`: schema to create the table in.

        `non_null_fields`: list of fields that are be NOT NULL

        `add_geom_column`: adds SELECT AddGeometryColumn POSTGIS function after each CREATE statement
        """

        with open(file_path, 'r') as json_file, open(output_path, 'w') as script:
            data = json.load(json_file)

            crs = data["crs_id"]
            for layer in data["Layers"]:
                table = layer["name"]
                geom = layer["geometry"]

                stmt = f"CREATE TABLE {schema}.{table} (\n"
                for field in layer["fields"]:
                    ((col, dtype),) = field.items()
                    stmt += f"   {col} {dtype}"
                    if col not in non_null_fields:
                        stmt += " NULL, \n"
                    else:
                        stmt += ",\n"

                stmt = stmt[:-3] + "\n)\n\n"
                
                if add_geom_column:
                    stmt += f"SELECT AddGeometryColumn('{schema}','{table}','geom',{crs},'{geom}',2);\n\n\n"

                script.write(stmt)

    def create_triggers(self, schema_select: str, schema_insert: str, output_path: str):
        """
        creates triggers for a database.

        selects all tables and their associative columns from schema name `schema_select`.

        writes a trigger for each table to insert into `schema_insert`.

        ...

        Arguments
        ---------

        `schema_select`: str

        `schema_insert`: str
        
        `output_path`: str
        """

        results = self.executer(
            f"""
            SELECT table_name, array_agg(CAST(column_name AS text)) AS fields 
            FROM information_schema.columns 
            WHERE table_schema = '{schema_select}' GROUP BY table_name""",
            True
        )

        func_logic = "BEGIN\n   INSERT INTO {schema}.{table} ({columns})\n   VALUES ({values});\n   RETURN NEW;\nEND;\n\n"

        with open(output_path, "w") as file:
            for row in results:
                (table, columns) = row
                file.write(
                    func_logic.format(
                        schema = schema_insert,
                        table=table, 
                        columns=",".join(columns),
                        values = ",".join(f"OLD.{col}" for col in columns)
                    )
                )
                      